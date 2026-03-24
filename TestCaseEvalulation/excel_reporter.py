"""
excel_reporter.py - Generate Excel reports from test results
"""
import json
import os
from datetime import datetime
from typing import Dict, List, Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart


# ── Color palette ──
COLORS = {
    "header_bg": "1E3A5F",
    "header_font": "FFFFFF",
    "pass_bg": "D6F5D6",
    "pass_font": "1B5E20",
    "fail_bg": "FDDEDE",
    "fail_font": "B71C1C",
    "invalid_bg": "FFF3CD",
    "invalid_font": "856404",
    "alt_row": "F5F8FF",
    "border": "BDD7EE",
    "subheader": "2E75B6",
    "subheader_font": "FFFFFF",
    "summary_bg": "EBF3FB",
    "total_bg": "1E3A5F",
    "total_font": "FFFFFF",
}

thin_border = Border(
    left=Side(style='thin', color=COLORS["border"]),
    right=Side(style='thin', color=COLORS["border"]),
    top=Side(style='thin', color=COLORS["border"]),
    bottom=Side(style='thin', color=COLORS["border"])
)


def _header_style(cell, text, bold=True, size=11):
    cell.value = text
    cell.font = Font(bold=bold, color=COLORS["header_font"], size=size, name="Arial")
    cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border


def _subheader_style(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=COLORS["subheader_font"], size=10, name="Arial")
    cell.fill = PatternFill("solid", fgColor=COLORS["subheader"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border


def _data_cell(cell, value, row_idx, status=None):
    cell.value = value
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = thin_border

    if status == "PASS":
        cell.fill = PatternFill("solid", fgColor=COLORS["pass_bg"])
    elif status == "FAIL":
        cell.fill = PatternFill("solid", fgColor=COLORS["fail_bg"])
    elif status == "INVALID":
        cell.fill = PatternFill("solid", fgColor=COLORS["invalid_bg"])
    elif row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])


def generate_excel_report(results: List[Dict[str, Any]], output_path: str) -> str:
    """Generate a comprehensive Excel report from test results."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Test Results ──
    ws_results = wb.active
    ws_results.title = "Test Results"
    _build_results_sheet(ws_results, results)

    # ── Sheet 2: Summary ──
    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, results, wb)

    # ── Sheet 3: Request Details ──
    ws_requests = wb.create_sheet("Request & Response Details")
    _build_request_sheet(ws_requests, results)

    wb.save(output_path)
    return output_path


def _build_results_sheet(ws, results: List[Dict]):
    """Build the main test results sheet."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"CEV Exemption Rules - Test Execution Report  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, color=COLORS["header_font"], size=13, name="Arial")
    title_cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = [
        "Test Case ID", "Test Case Name", "Rule Tested", "Test Type",
        "Description", "Expected Status", "Expected Exemption",
        "Expected Reason", "Expected Rule Fired",
        "Actual Status", "Actual Exemption", "Actual Reason",
        "Actual Rule Fired", "Verification Status", "Failure Reason"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        _header_style(cell, header)

    ws.row_dimensions[2].height = 35

    # Data rows
    for row_idx, result in enumerate(results, 3):
        status = result.get("verification_status", "")

        data = [
            result.get("test_case_id", ""),
            result.get("test_case_name", ""),
            result.get("rule_being_tested", ""),
            result.get("test_type", ""),
            result.get("description", ""),
            result.get("expected_status", "SUCCESS"),
            result.get("expected_exemption_status", ""),
            result.get("expected_exemption_reason", ""),
            result.get("expected_rule_fired", ""),
            result.get("actual_status", ""),
            result.get("actual_exemption_status", ""),
            result.get("actual_exemption_reason", ""),
            result.get("actual_rule_fired", ""),
            status,
            result.get("failure_reason", "")
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            col_status = status if col >= 14 else None
            _data_cell(cell, value, row_idx, col_status)

            # Bold the status cell
            if col == 14:
                cell.font = Font(
                    bold=True,
                    name="Arial",
                    size=9,
                    color=COLORS["pass_font"] if status == "PASS"
                    else COLORS["fail_font"] if status == "FAIL"
                    else COLORS["invalid_font"]
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 28

    # Column widths
    col_widths = [12, 30, 25, 12, 40, 14, 16, 30, 25, 14, 16, 30, 25, 16, 35]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(results) + 2}"


def _build_summary_sheet(ws, results: List[Dict], wb):
    """Build summary dashboard sheet."""
    ws.sheet_view.showGridLines = False

    total = len(results)
    passed = sum(1 for r in results if r.get("verification_status") == "PASS")
    failed = sum(1 for r in results if r.get("verification_status") == "FAIL")
    invalid = sum(1 for r in results if r.get("verification_status") == "INVALID")
    pass_rate = (passed / total * 100) if total else 0

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "TEST EXECUTION SUMMARY DASHBOARD"
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Summary cards
    cards = [
        ("Total Test Cases", total, "2E75B6", "FFFFFF"),
        ("✅ Passed", passed, "1B5E20", "FFFFFF"),
        ("❌ Failed", failed, "B71C1C", "FFFFFF"),
        ("⚠️ Invalid Data", invalid, "856404", "FFFFFF"),
        ("Pass Rate", f"{pass_rate:.1f}%", "1E3A5F", "FFFFFF"),
    ]

    for i, (label, value, bg, fg) in enumerate(cards):
        col = i * 2 + 1
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)

        label_cell = ws.cell(row=3, column=col, value=label)
        label_cell.font = Font(bold=True, color=fg, size=10, name="Arial")
        label_cell.fill = PatternFill("solid", fgColor=bg)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        val_cell = ws.cell(row=4, column=col, value=value)
        val_cell.font = Font(bold=True, color=fg, size=22, name="Arial")
        val_cell.fill = PatternFill("solid", fgColor=bg)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=5, column=col + 1).fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 40
        ws.row_dimensions[5].height = 10

    # Breakdown by Test Type
    ws.merge_cells("A7:D7")
    h = ws["A7"]
    h.value = "Breakdown by Test Type"
    h.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    h.fill = PatternFill("solid", fgColor=COLORS["subheader"])
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 25

    type_headers = ["Test Type", "Total", "Pass", "Fail"]
    for col, hdr in enumerate(type_headers, 1):
        c = ws.cell(row=8, column=col, value=hdr)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    type_stats = {}
    for r in results:
        tt = r.get("test_type", "unknown")
        if tt not in type_stats:
            type_stats[tt] = {"total": 0, "pass": 0, "fail": 0}
        type_stats[tt]["total"] += 1
        vs = r.get("verification_status", "")
        if vs == "PASS":
            type_stats[tt]["pass"] += 1
        else:
            type_stats[tt]["fail"] += 1

    for row_i, (tt, stats) in enumerate(type_stats.items(), 9):
        row_data = [tt, stats["total"], stats["pass"], stats["fail"]]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor="F5F8FF" if row_i % 2 == 0 else "FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

    # Breakdown by Rule
    rule_row_start = max(10, 9 + len(type_stats) + 1)
    ws.merge_cells(f"A{rule_row_start}:F{rule_row_start}")
    h2 = ws[f"A{rule_row_start}"]
    h2.value = "Breakdown by Rule"
    h2.font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    h2.fill = PatternFill("solid", fgColor=COLORS["subheader"])
    h2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[rule_row_start].height = 25

    rule_headers = ["Rule", "Total", "Pass", "Fail", "Invalid", "Pass Rate"]
    for col, hdr in enumerate(rule_headers, 1):
        c = ws.cell(row=rule_row_start + 1, column=col, value=hdr)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    rule_stats = {}
    for r in results:
        rule = r.get("rule_being_tested", "unknown")
        if rule not in rule_stats:
            rule_stats[rule] = {"total": 0, "pass": 0, "fail": 0, "invalid": 0}
        rule_stats[rule]["total"] += 1
        vs = r.get("verification_status", "")
        if vs == "PASS":
            rule_stats[rule]["pass"] += 1
        elif vs == "FAIL":
            rule_stats[rule]["fail"] += 1
        else:
            rule_stats[rule]["invalid"] += 1

    for row_i, (rule, stats) in enumerate(rule_stats.items(), rule_row_start + 2):
        rate = f"{stats['pass'] / stats['total'] * 100:.0f}%" if stats["total"] else "0%"
        row_data = [rule, stats["total"], stats["pass"], stats["fail"], stats["invalid"], rate]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor="F5F8FF" if row_i % 2 == 0 else "FFFFFF")
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

    # Column widths
    for col, w in enumerate([18, 10, 10, 10, 10, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_request_sheet(ws, results: List[Dict]):
    """Build request/response details sheet."""
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Request & Response Details"
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    c.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Test Case ID", "Test Case Name", "Verification Status", "Request JSON", "Response JSON"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        _header_style(cell, hdr)
    ws.row_dimensions[2].height = 30

    for row_idx, result in enumerate(results, 3):
        status = result.get("verification_status", "")
        data = [
            result.get("test_case_id", ""),
            result.get("test_case_name", ""),
            status,
            json.dumps(result.get("request_json", {}), indent=2),
            json.dumps(result.get("response_json", {}), indent=2),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            _data_cell(cell, value, row_idx, status if col == 3 else None)
            if col in [4, 5]:
                cell.font = Font(name="Courier New", size=8)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[row_idx].height = 80

    col_widths = [14, 30, 16, 50, 50]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def compute_summary_stats(results: List[Dict]) -> Dict[str, Any]:
    """Compute summary statistics from results."""
    total = len(results)
    passed = sum(1 for r in results if r.get("verification_status") == "PASS")
    failed = sum(1 for r in results if r.get("verification_status") == "FAIL")
    invalid = sum(1 for r in results if r.get("verification_status") == "INVALID")
    pass_rate = (passed / total * 100) if total else 0

    by_type = {}
    by_rule = {}
    for r in results:
        tt = r.get("test_type", "unknown")
        rule = r.get("rule_being_tested", "unknown")
        vs = r.get("verification_status", "")

        if tt not in by_type:
            by_type[tt] = {"total": 0, "pass": 0, "fail": 0, "invalid": 0}
        by_type[tt]["total"] += 1
        by_type[tt][vs.lower() if vs.lower() in ["pass", "fail", "invalid"] else "invalid"] += 1

        if rule not in by_rule:
            by_rule[rule] = {"total": 0, "pass": 0, "fail": 0, "invalid": 0}
        by_rule[rule]["total"] += 1
        by_rule[rule][vs.lower() if vs.lower() in ["pass", "fail", "invalid"] else "invalid"] += 1

    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "invalid": invalid,
        "pass_rate": pass_rate,
        "by_type": by_type,
        "by_rule": by_rule,
    }
